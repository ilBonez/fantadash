#!/usr/bin/env python3
"""Converte il workbook Fantacalcio Classic in src/data/listone.json.

Uso:
    python scripts/ingest_xlsx.py [percorso.xlsx]

Senza argomenti prende il .xlsx piu recente in data/.

Il workbook e' l'unica fonte di verita' della dashboard: listone per reparto,
matrice degli abbinamenti di calendario, coppie e terzetti consigliati,
infortunati, statistiche 2025/26 e parametri di lega finiscono tutti nel JSON
che l'app importa. Non ci sono overlay da agganciare a mano ne' fetch di rete.

L'unica eccezione e' data/ceduti.txt: l'elenco di chi ha lasciato la Serie A
dopo la data del workbook, che va tolto dal listone senza aspettare un file
nuovo. Non aggiunge dati, ne toglie.
"""
from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "src" / "data" / "listone.json"
CEDUTI = ROOT / "data" / "ceduti.txt"

# Nei fogli di reparto la riga 1 e' il titolo, la 2 il conteggio, la 3 l'header.
HEADER_ROW = 3

SHEETS = {"Portieri": "P", "Difensori": "D", "Centrocampisti": "C", "Attaccanti": "A"}

# Ordine delle fasce dal workbook: serve per ordinare e colorare.
FASCE = ["Top", "1a fascia", "2a fascia", "3a fascia", "4a fascia", "Scommessa"]

ORDINALE = re.compile(r"^([1-9])o$")


def pick_source(argv: list[str]) -> Path:
    if len(argv) > 1:
        return Path(argv[1])
    candidates = sorted((ROOT / "data").glob("*.xlsx"), key=lambda p: p.stat().st_mtime)
    if not candidates:
        sys.exit("Nessun .xlsx trovato in data/. Passa il percorso come argomento.")
    return candidates[-1]


def txt(v) -> str:
    """Testo pulito. Il workbook usa '-' come 'nessun valore'."""
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s == "-" else s


def num(v):
    """Numero, oppure None se la cella e' vuota o non numerica."""
    if v is None or v == "" or v == "-":
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return int(f) if f == int(f) else round(f, 3)


def num0(v) -> float:
    n = num(v)
    return 0 if n is None else n


def ordinale(v) -> int | None:
    """'1o' -> 1. Rigorista e tiratore da fermo sono scritti cosi."""
    m = ORDINALE.match(txt(v))
    return int(m.group(1)) if m else None


def header_of(ws) -> list[str]:
    row = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
    return [str(c).strip() if c is not None else "" for c in row]


# --- listone per reparto ----------------------------------------------------

# Colonne attese nei quattro fogli di reparto. Se il workbook cambia intestazioni
# lo script si ferma qui invece di produrre un JSON silenziosamente vuoto.
COLS = [
    "Prio", "Indice", "Giocatore", "Squadra", "Ruolo Mantra", "Qt.I", "Qt.A", "FVM",
    "Prezzo consigliato", "Prezzo max", "Fascia", "Gerarchia stimata", "NOTA",
    "Rigorista", "Calci piazzati",
    "PG 25/26", "MV 25/26", "FM 25/26", "Gol 25/26", "Assist 25/26",
    "Rig. segnati 25/26", "Amm. 25/26", "Esp. 25/26",
    "PG 26/27 (su 2)", "MV 26/27", "FM 26/27", "Gol 26/27", "Assist 26/27",
    "Squadra 25/26", "Nuovo/Trasf.", "FM ponderata",
    "Miglior abbinamento (calendario)", "Trasferte in comune", "Rank FVM di ruolo",
    "Dettaglio infortunio / stato",
]


def leggi_reparto(ws, ruolo: str) -> list[dict]:
    header = header_of(ws)
    missing = [c for c in COLS if c not in header]
    if missing:
        sys.exit(f"Foglio '{ws.title}': colonne mancanti {missing}")
    idx = {c: header.index(c) for c in COLS}
    g = lambda row, c: row[idx[c]]  # noqa: E731

    out = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        nome = txt(g(row, "Giocatore"))
        if not nome:
            continue

        pg25 = num(g(row, "PG 25/26"))
        s25 = None
        if pg25 is not None:
            s25 = {
                "pg": pg25,
                "mv": num(g(row, "MV 25/26")),
                "fm": num(g(row, "FM 25/26")),
                "gol": num0(g(row, "Gol 25/26")),
                "ass": num0(g(row, "Assist 25/26")),
                "rig": txt(g(row, "Rig. segnati 25/26")) or "0/0",
                "amm": num0(g(row, "Amm. 25/26")),
                "esp": num0(g(row, "Esp. 25/26")),
                "squadra": txt(g(row, "Squadra 25/26")),
            }

        pg26 = num0(g(row, "PG 26/27 (su 2)"))
        s26 = {
            "pg": pg26,
            "mv": num(g(row, "MV 26/27")),
            "fm": num(g(row, "FM 26/27")),
            "gol": num0(g(row, "Gol 26/27")),
            "ass": num0(g(row, "Assist 26/27")),
        }

        out.append({
            "r": ruolo,
            "nome": nome,
            "squadra": txt(g(row, "Squadra")),
            "rm": txt(g(row, "Ruolo Mantra")),
            "prio": num0(g(row, "Prio")),
            "indice": num0(g(row, "Indice")),
            "qtI": num0(g(row, "Qt.I")),
            "qtA": num0(g(row, "Qt.A")),
            "fvm": num0(g(row, "FVM")),
            "cons": num0(g(row, "Prezzo consigliato")),
            "max": num0(g(row, "Prezzo max")),
            "fascia": txt(g(row, "Fascia")),
            "gerarchia": txt(g(row, "Gerarchia stimata")),
            "nota": txt(g(row, "NOTA")),
            "rig": ordinale(g(row, "Rigorista")),
            "piaz": ordinale(g(row, "Calci piazzati")),
            "fmPond": num0(g(row, "FM ponderata")),
            "nuovo": txt(g(row, "Nuovo/Trasf.")).lower().startswith("s"),
            "abb": txt(g(row, "Miglior abbinamento (calendario)")),
            "abbTras": num0(g(row, "Trasferte in comune")),
            "rankFvm": num0(g(row, "Rank FVM di ruolo")),
            "infDettaglio": txt(g(row, "Dettaglio infortunio / stato")),
            "s25": s25,
            "s26": s26,
        })
    return out


# --- foglio DB: da' la sigla di ogni giocatore e quindi la chiave ------------


def leggi_db(ws) -> dict[tuple[str, str], str]:
    """(nome, squadra) -> sigla squadra, letta dal foglio DB."""
    header = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(values_only=True))]
    i_nome, i_cod, i_sq = header.index("Nome"), header.index("Cod"), header.index("Squadra")
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome, cod, sq = txt(row[i_nome]), txt(row[i_cod]), txt(row[i_sq])
        if nome and cod:
            out[(nome, sq)] = cod
    return out


# --- abbinamenti di calendario ----------------------------------------------


def trova_riga(ws, prefisso: str) -> int | None:
    for i, row in enumerate(ws.iter_rows(max_col=1, values_only=True), 1):
        if row[0] and str(row[0]).startswith(prefisso):
            return i
    return None


def leggi_matrice(ws) -> tuple[list[str], dict[str, dict[str, int]]]:
    """Matrice trasferte in comune: squadra -> squadra -> giornate su 38."""
    r0 = trova_riga(ws, "MATRICE COMPLETA")
    if r0 is None:
        sys.exit("Foglio Abbinamenti: manca la sezione MATRICE COMPLETA")
    intestazione = [txt(c.value) for c in ws[r0 + 2]]
    squadre = [s for s in intestazione[1:] if s]

    matrice: dict[str, dict[str, int]] = {}
    for i in range(len(squadre)):
        riga = ws[r0 + 3 + i]
        nome = txt(riga[0].value)
        if not nome:
            break
        matrice[nome] = {}
        for j, avversaria in enumerate(squadre):
            if avversaria == nome:
                continue
            v = num(riga[j + 1].value)
            if v is not None:
                matrice[nome][avversaria] = int(v)
    return squadre, matrice


def leggi_coppie(ws, prefisso: str) -> list[dict]:
    r0 = trova_riga(ws, prefisso)
    if r0 is None:
        return []
    out = []
    for row in ws.iter_rows(min_row=r0 + 2, values_only=True):
        if not row[1]:
            break
        out.append({
            "a": txt(row[1]),
            "b": txt(row[2]),
            "t": int(num0(row[3])),
            "giudizio": txt(row[4]),
        })
    return out


def leggi_terzetti(ws) -> list[dict]:
    r0 = trova_riga(ws, "TOP 20 TERZETTI")
    if r0 is None:
        return []
    out = []
    for row in ws.iter_rows(min_row=r0 + 2, values_only=True):
        if not row[1]:
            break
        out.append({
            "p": [txt(row[1]), txt(row[2]), txt(row[3])],
            "t": [int(num0(row[4])), int(num0(row[5])), int(num0(row[6]))],
            "tot": int(num0(row[7])),
            "costo": int(num0(row[8])),
            "indice": num0(row[9]),
        })
    return out


# --- infortunati -------------------------------------------------------------


def leggi_infortunati(ws) -> dict[tuple[str, str], dict]:
    header = header_of(ws) if txt(ws["A3"].value) == "Giocatore" else None
    if header is None:
        return {}
    i = {c: header.index(c) for c in ("Giocatore", "Squadra", "Stato", "Dettaglio e tempi di recupero")}
    out = {}
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        nome, squadra = txt(row[i["Giocatore"]]), txt(row[i["Squadra"]])
        if not nome:
            continue
        out[(nome, squadra)] = {
            "stato": txt(row[i["Stato"]]),
            "dettaglio": txt(row[i["Dettaglio e tempi di recupero"]]),
        }
    return out


# --- giocatori usciti dopo la data del workbook ------------------------------


def leggi_ceduti() -> list[str]:
    """Le chiavi da togliere dal listone, da data/ceduti.txt."""
    if not CEDUTI.exists():
        return []
    righe = []
    for riga in CEDUTI.read_text(encoding="utf-8").splitlines():
        riga = riga.split("#", 1)[0].strip()
        if riga:
            righe.append(riga)
    return righe


# --- parametri di lega dal foglio Guida --------------------------------------


def leggi_parametri(ws) -> dict:
    """Budget, squadre e slot per reparto: la taratura con cui e' stato costruito
    il listone. La dashboard la usa come default delle impostazioni."""
    valori: dict[str, float] = {}
    for row in ws.iter_rows(values_only=True):
        etichetta = txt(row[0])
        if etichetta and len(row) > 1:
            v = num(row[1])
            if v is not None:
                valori.setdefault(etichetta, v)

    slots, quota = {}, {}
    for row in ws.iter_rows(values_only=True):
        reparto = txt(row[0])
        if reparto in SHEETS and len(row) > 2:
            r = SHEETS[reparto]
            slots[r] = int(num0(row[1]))
            quota[r] = num0(row[2])

    return {
        "budget": int(valori.get("Budget per squadra (crediti)", 500)),
        "squadre": int(valori.get("Numero di squadre nella lega", 10)),
        "slots": slots or {"P": 3, "D": 8, "C": 8, "A": 6},
        "quotaReparto": quota,
        "compressione": valori.get("Compressione della curva (0,5 - 1,0)"),
    }


# Un titolo di sezione nella Guida occupa da solo la riga e comincia in
# maiuscolo: "NOTE SUL METODO", "LA COLONNA NOTA (aggiornata al ...)".
TITOLO_SEZIONE = re.compile(r"^[A-Z][A-Z0-9 ,'()/-]{3,}")


def leggi_guida(ws) -> dict:
    """Le sezioni discorsive della Guida: legenda delle note e note sul metodo."""
    sezioni: dict[str, list[list[str]]] = {}
    corrente: str | None = None
    for row in ws.iter_rows(values_only=True):
        celle = [txt(c) for c in row]
        prima = celle[0] if celle else ""
        resto = [c for c in celle[1:] if c]
        if prima and not resto and TITOLO_SEZIONE.match(prima):
            corrente = prima
            sezioni[corrente] = []
        elif corrente and prima and resto:
            sezioni[corrente].append([prima, " ".join(resto)])

    def sezione(prefisso: str) -> list[list[str]]:
        for titolo, voci in sezioni.items():
            if titolo.startswith(prefisso):
                return voci
        return []

    return {
        "legendaNota": sezione("LA COLONNA NOTA"),
        "metodo": sezione("NOTE SUL METODO"),
    }


# --- main --------------------------------------------------------------------


def main() -> None:
    src = pick_source(sys.argv)
    wb = openpyxl.load_workbook(src, data_only=True)

    giocatori: list[dict] = []
    for foglio, ruolo in SHEETS.items():
        giocatori.extend(leggi_reparto(wb[foglio], ruolo))

    cod_di = leggi_db(wb["DB"]) if "DB" in wb.sheetnames else {}
    infortuni = leggi_infortunati(wb["Infortunati"]) if "Infortunati" in wb.sheetnames else {}

    # Sigle squadra: dal foglio DB, con ripiego sulle prime tre lettere.
    sigle: dict[str, str] = {}
    for (_, squadra), cod in cod_di.items():
        sigle.setdefault(squadra, cod)

    # Gli id si assegnano PRIMA di togliere i ceduti: cosi' aggiungere un nome a
    # data/ceduti.txt non rinumera tutti gli altri, e le assegnazioni gia' fatte
    # in dashboard continuano a puntare al giocatore giusto.
    orfani: list[str] = []
    for i, p in enumerate(sorted(giocatori, key=lambda x: (x["r"], x["prio"], x["nome"])), 1):
        cod = cod_di.get((p["nome"], p["squadra"])) or sigle.get(p["squadra"]) or p["squadra"][:3].upper()
        p["id"] = i
        p["cod"] = cod
        # Stessa chiave usata dal workbook per gli abbinamenti: "Svilar (ROM)".
        p["chiave"] = f"{p['nome']} ({cod})"
        p["fasciaIdx"] = FASCE.index(p["fascia"]) if p["fascia"] in FASCE else len(FASCE)

        inf = infortuni.get((p["nome"], p["squadra"]))
        if inf:
            p["inf"] = inf
        elif p["infDettaglio"]:
            p["inf"] = {"stato": "", "dettaglio": p["infDettaglio"]}
        p.pop("infDettaglio", None)

        if (p["nome"], p["squadra"]) not in cod_di:
            orfani.append(f"{p['nome']} ({p['squadra']})")

    # Chi ha lasciato la Serie A dopo la data del workbook esce dal listone.
    ceduti = leggi_ceduti()
    if ceduti:
        presenti = {p["chiave"] for p in giocatori}
        ignoti = [c for c in ceduti if c not in presenti]
        if ignoti:
            sys.exit(
                f"data/ceduti.txt: chiavi non trovate nel listone {ignoti}. "
                "Il formato e' \"Cognome (SIG)\", come nella colonna Chiave del foglio DB."
            )
        via = set(ceduti)
        giocatori = [p for p in giocatori if p["chiave"] not in via]

    giocatori.sort(key=lambda p: p["id"])
    per_chiave = {p["chiave"]: p["id"] for p in giocatori}

    # L'abbinamento del workbook e' una chiave testuale: la risolviamo in id.
    # Se punta a un ceduto lo si azzera, altrimenti la dashboard consiglierebbe
    # di accoppiarsi a un giocatore che non esiste piu'.
    abb_persi = []
    for p in giocatori:
        p["abbId"] = per_chiave.get(p["abb"])
        if p["abb"] and p["abbId"] is None:
            abb_persi.append(f"{p['chiave']} -> {p['abb']}")
            p["abb"] = ""
            p["abbTras"] = 0

    ws_abb = wb["Abbinamenti"]
    squadre, matrice = leggi_matrice(ws_abb)
    payload = {
        "stagione": txt(wb["Guida"]["A1"].value),
        "descrizione": txt(wb["Guida"]["A2"].value),
        "sorgente": src.name,
        "generatoIl": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "parametri": leggi_parametri(wb["Guida"]),
        "guida": leggi_guida(wb["Guida"]),
        "conteggi": {r: sum(1 for p in giocatori if p["r"] == r) for r in SHEETS.values()},
        "fasce": FASCE,
        "squadre": squadre,
        "sigle": sigle,
        "matrice": matrice,
        "coppieMigliori": leggi_coppie(ws_abb, "MIGLIORI COPPIE"),
        "coppiePeggiori": leggi_coppie(ws_abb, "PEGGIORI COPPIE"),
        "terzettiPortieri": leggi_terzetti(ws_abb),
        "comeSiLegge": txt(ws_abb.cell(trova_riga(ws_abb, "Come si legge") or 1, 1).value),
        "giocatori": giocatori,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")

    print(f"{src.name} -> {OUT.relative_to(ROOT)}")
    print(f"  {len(giocatori)} giocatori {payload['conteggi']}")
    print(f"  {len(squadre)} squadre, {len(payload['coppieMigliori'])} coppie top, "
          f"{len(payload['terzettiPortieri'])} terzetti portieri")
    print(f"  {sum(1 for p in giocatori if 'inf' in p)} con nota infortunio")
    if ceduti:
        print(f"  {len(ceduti)} tolti da data/ceduti.txt: {', '.join(ceduti)}")
    if orfani:
        print(f"  {len(orfani)} giocatori senza sigla nel foglio DB: {', '.join(orfani[:5])}")
    if abb_persi:
        print(f"  {len(abb_persi)} abbinamenti azzerati (puntavano a un ceduto): "
              f"{', '.join(abb_persi[:5])}")


if __name__ == "__main__":
    main()
